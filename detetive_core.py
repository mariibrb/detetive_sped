"""Núcleo do Detetive (sem UI) — usado pela app Django `detetive_fiscal` (modulo Sentinela)."""

import io
import re
import uuid
import zipfile
from collections import Counter
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
from xlsxwriter.utility import xl_col_to_name, xl_rowcol_to_cell

ROTULO_CLIENTE = "SPED Cliente"
ROTULO_NASCEL = "SPED Nascel"
SUF_CLIENTE = "_Cliente"
SUF_NASCEL = "_Nascel"

CHAVES_CANDIDATAS = [
    "CHV_NFE",
    "CHV_CTE",
    "CHV_NFSE",
    "CHAVE_ACESSO",
    "CHV_DOCe",
    "CNPJ",
    "COD_PART",
]

ALIASES_POR_REG = {
    "C100": [("C09", "CHV_NFE")],
    "0150": [("C05", "CNPJ"), ("C02", "COD_PART")],
}


def _norm_val(v):
    if pd.isna(v) or v is None:
        return ""
    return str(v).strip()


def _ref_registro_sped_row(row: pd.Series) -> str:
    for base in ("CHV_NFE", "NUM_DOC", "CNPJ", "COD_PART", "CHAVE_ACESSO", "CHV_CTE"):
        for suf in ("", SUF_CLIENTE, SUF_NASCEL):
            col = f"{base}{suf}" if suf else base
            if col in row.index:
                v = _norm_val(row[col])
                if v and v.lower() not in ("nan", "none"):
                    return f"{base} {v}"
    return "—"


def _detalhe_divergencia_campos(row: pd.Series) -> str:
    raw = row.get("COLUNAS_DIVERGENTES", "")
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    partes = []
    for col in str(raw).split(","):
        col = col.strip()
        if not col:
            continue
        if "(" in col:
            partes.append(col)
            continue
        c1, c2 = f"{col}{SUF_CLIENTE}", f"{col}{SUF_NASCEL}"
        if c1 in row.index and c2 in row.index:
            v1, v2 = _norm_val(row[c1]), _norm_val(row[c2])
            if v1 != v2:
                partes.append(f"{col}: Cliente «{v1}» | Nascel «{v2}»")
            else:
                partes.append(col)
        else:
            partes.append(col)
    return "; ".join(partes) if partes else str(raw)


def montar_onde_agir_sped(
    out_div: pd.DataFrame,
    out_so_c: pd.DataFrame,
    out_so_n: pd.DataFrame,
    max_div_detalhe: int = 500,
) -> pd.DataFrame:
    linhas: list[dict] = []
    n = 0

    if out_so_c is not None and not out_so_c.empty and "ABA_SPED" in out_so_c.columns:
        for aba, grp in out_so_c.groupby("ABA_SPED"):
            n += 1
            linhas.append(
                {
                    "#": n,
                    "Prioridade": "Média",
                    "Tipo": f"Só no {ROTULO_CLIENTE}",
                    "Referencia": str(aba),
                    "Acao": f"{len(grp)} linha(s) só no Cliente neste registro. Abrir aba **So_Cliente** (filtrar ABA_SPED = {aba}).",
                }
            )

    if out_so_n is not None and not out_so_n.empty and "ABA_SPED" in out_so_n.columns:
        for aba, grp in out_so_n.groupby("ABA_SPED"):
            n += 1
            linhas.append(
                {
                    "#": n,
                    "Prioridade": "Média",
                    "Tipo": f"Só no {ROTULO_NASCEL}",
                    "Referencia": str(aba),
                    "Acao": f"{len(grp)} linha(s) só na Nascel neste registro. Abrir aba **So_Nascel** (filtro ABA_SPED = {aba}).",
                }
            )

    if out_div is not None and not out_div.empty:
        for i, (_, row) in enumerate(out_div.iterrows()):
            if i >= max_div_detalhe:
                break
            n += 1
            aba = row.get("ABA_SPED", "")
            ref = _ref_registro_sped_row(row)
            det = _detalhe_divergencia_campos(row)
            linhas.append(
                {
                    "#": n,
                    "Prioridade": "Alta",
                    "Tipo": "Mesma chave, campos diferentes",
                    "Referencia": f"{aba} · {ref}",
                    "Acao": det or _norm_val(row.get("COLUNAS_DIVERGENTES", "")),
                }
            )
        overflow = len(out_div) - min(len(out_div), max_div_detalhe)
        if overflow > 0:
            n += 1
            linhas.append(
                {
                    "#": n,
                    "Prioridade": "Alta",
                    "Tipo": "Limite da lista",
                    "Referencia": "—",
                    "Acao": f"Mais {overflow} divergência(s) na aba **Campos_divergentes** (mesmo detalhe por linha).",
                }
            )

    if not linhas:
        linhas.append(
            {
                "#": 1,
                "Prioridade": "—",
                "Tipo": "Sem pendências",
                "Referencia": "—",
                "Acao": "Nada listado: nenhuma diferença nos blocos comparados (ou filtros vazios).",
            }
        )

    return pd.DataFrame(linhas)


# --- Somatório por CFOP (C100 amarra C190) — comparar dois SPEDs ---

_COLS_C190_SOMA = ["VL_OPR", "VL_BC_ICMS", "VL_ICMS", "VL_BC_ICMS_ST", "VL_ICMS_ST", "VL_IPI"]

_COL_CFOP_SOM = "CFOP_SPED_soma"

# --- Rótulos em português no Excel (colunas identificáveis; chave de acesso em destaque) ---
_LABEL_ORIGEM_SPED = "Origem do ficheiro (SPED)"
_LABEL_CHAVE_NFE = "Chave de acesso da NF-e (44 dígitos)"
_LABEL_CHAVE_CTE = "Chave de acesso CT-e / documento (44 dígitos)"
_LABEL_CHAVE_ONDE_AGIR = "Chave de acesso da nota / doc. (44 dígitos)"
_LABEL_CFOP_SOM_TXT = "CFOP em texto (usado no somatório / tabela dinâmica)"

_SOMA_COL_PT = {
    "VL_OPR": "Soma — valor da operação (VL_OPR)",
    "VL_BC_ICMS": "Soma — base de cálculo ICMS (VL_BC_ICMS)",
    "VL_ICMS": "Soma — valor do ICMS (VL_ICMS)",
    "VL_BC_ICMS_ST": "Soma — base ICMS ST (VL_BC_ICMS_ST)",
    "VL_ICMS_ST": "Soma — valor ICMS ST (VL_ICMS_ST)",
    "VL_IPI": "Soma — valor do IPI (VL_IPI)",
}

_CAB_CFOP_EXCEL = ["CFOP"] + [_SOMA_COL_PT[c] for c in _COLS_C190_SOMA]

_COL_VALOR_PT = {
    "VL_OPR": "Valor da operação (VL_OPR)",
    "VL_BC_ICMS": "Base de cálculo ICMS (VL_BC_ICMS)",
    "VL_ICMS": "Valor do ICMS (VL_ICMS)",
    "VL_BC_ICMS_ST": "Base ICMS ST (VL_BC_ICMS_ST)",
    "VL_ICMS_ST": "Valor ICMS ST (VL_ICMS_ST)",
    "VL_IPI": "Valor do IPI (VL_IPI)",
    "VL_RED_BC": "Valor redução da base de cálculo (VL_RED_BC)",
    "VL_DOC": "Valor do documento (VL_DOC)",
}

_COL_DETALHE_C190 = {
    "LADO": _LABEL_ORIGEM_SPED,
    "CHV_NFE": _LABEL_CHAVE_NFE,
    "CHAVE_ACESSO": _LABEL_CHAVE_NFE,
    "CFOP": "CFOP",
    _COL_CFOP_SOM: _LABEL_CFOP_SOM_TXT,
    "NUM_DOC": "Número do documento",
    "SER": "Série",
    "COD_PART": "Código do participante",
    "CNPJ": "CNPJ",
    "CST_ICMS": "CST ICMS",
    "REG": "Registro SPED",
}
for _k, _v in _COL_VALOR_PT.items():
    _COL_DETALHE_C190.setdefault(_k, _v)

_COL_DETALHE_D190 = {
    "LADO": _LABEL_ORIGEM_SPED,
    "CHV_DOC": _LABEL_CHAVE_CTE,
    "CHV_CTE": _LABEL_CHAVE_CTE,
    "CHV_NFE": _LABEL_CHAVE_CTE,
    "CFOP": "CFOP",
    _COL_CFOP_SOM: _LABEL_CFOP_SOM_TXT,
    "NUM_DOC": "Número do documento",
    "SER": "Série",
    "COD_PART": "Código do participante",
    "CNPJ": "CNPJ",
    "CST_ICMS": "CST ICMS",
    "REG": "Registro SPED",
}
for _k, _v in _COL_VALOR_PT.items():
    _COL_DETALHE_D190.setdefault(_k, _v)

_COL_C100_PLAN = {
    "LADO": _LABEL_ORIGEM_SPED,
    "CHV_NFE": _LABEL_CHAVE_NFE,
    "CHAVE_ACESSO": _LABEL_CHAVE_NFE,
    "NUM_DOC": "Número do documento",
    "SER": "Série",
    "COD_PART": "Código do participante",
    "CNPJ": "CNPJ",
    "VL_DOC": _COL_VALOR_PT["VL_DOC"],
    "IND_OPER": "Indicador operação",
    "IND_EMIT": "Indicador emitente",
    "COD_MOD": "Modelo documento",
    "REG": "Registro SPED",
}


def _col_excel_idx(df: pd.DataFrame, interno: str, mapa: dict[str, str]) -> int | None:
    cols = list(df.columns)
    lab = mapa.get(interno, interno)
    if lab in cols:
        return cols.index(lab)
    if interno in cols:
        return cols.index(interno)
    return None


def _xlsx_largura_coluna_chave(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
    mapa: dict[str, str],
    largura: float = 50,
) -> None:
    """Garante coluna legível para chave de 44 dígitos (NF-e ou CT-e)."""
    if sheet_name not in writer.sheets or df is None or df.empty:
        return
    ws = writer.sheets[sheet_name]
    for interno in ("CHV_NFE", "CHV_DOC"):
        j = _col_excel_idx(df, interno, mapa)
        if j is not None:
            ws.set_column(j, j, largura)
            return


def _df_export_c190_planilha(df: pd.DataFrame, lado: str) -> pd.DataFrame:
    """Detalhe C190 com cabeçalhos em português; chave NF-e logo após a origem."""
    x = _df_export_sped_com_cfop_soma(df, lado)
    if x.empty:
        return x
    out = x.copy()
    if "CHV_NFE" not in out.columns:
        out["CHV_NFE"] = ""
    if "CHAVE_ACESSO" in x.columns:
        out["CHV_NFE"] = out["CHV_NFE"].replace("", pd.NA).fillna(x["CHAVE_ACESSO"])
    out["CHV_NFE"] = out["CHV_NFE"].map(
        lambda v: _digits_chave(v) if pd.notna(v) and str(v).strip() else ""
    )
    pref = ["LADO", "CHV_NFE", "CFOP", _COL_CFOP_SOM, "CST_ICMS"]
    vl_rest = [c for c in _COLS_C190_SOMA if c in out.columns and c not in pref]
    outros = [c for c in out.columns if c not in pref + vl_rest]
    ordered = [c for c in pref if c in out.columns] + vl_rest + outros
    out = out[ordered]
    ren = {k: v for k, v in _COL_DETALHE_C190.items() if k in out.columns}
    return out.rename(columns=ren)


def _df_export_d190_planilha(df: pd.DataFrame, lado: str) -> pd.DataFrame:
    x = _df_export_sped_com_cfop_soma(df, lado)
    if x.empty:
        return x
    out = x.copy()
    if "CHV_DOC" not in out.columns:
        out["CHV_DOC"] = ""
    if "CHV_CTE" in x.columns:
        out["CHV_DOC"] = out["CHV_DOC"].replace("", pd.NA).fillna(x["CHV_CTE"])
    if "CHV_NFE" in x.columns and out["CHV_DOC"].eq("").all():
        out["CHV_DOC"] = out["CHV_DOC"].mask(out["CHV_DOC"].eq(""), x["CHV_NFE"])
    out["CHV_DOC"] = out["CHV_DOC"].map(
        lambda v: _digits_chave(v) if pd.notna(v) and str(v).strip() else ""
    )
    pref = ["LADO", "CHV_DOC", "CFOP", _COL_CFOP_SOM, "CST_ICMS"]
    vl_rest = [c for c in _COLS_C190_SOMA if c in out.columns and c not in pref]
    if "VL_RED_BC" in out.columns and "VL_RED_BC" not in pref + vl_rest:
        vl_rest.append("VL_RED_BC")
    outros = [c for c in out.columns if c not in pref + vl_rest]
    ordered = [c for c in pref if c in out.columns] + vl_rest + outros
    out = out[ordered]
    ren = {k: v for k, v in _COL_DETALHE_D190.items() if k in out.columns}
    return out.rename(columns=ren)


def _df_c100_planilha(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df if df is not None else pd.DataFrame()
    out = df.copy()
    if "CHV_NFE" not in out.columns:
        out["CHV_NFE"] = ""
    if "CHAVE_ACESSO" in df.columns:
        out["CHV_NFE"] = out["CHV_NFE"].replace("", pd.NA).fillna(df["CHAVE_ACESSO"])
    out["CHV_NFE"] = out["CHV_NFE"].map(
        lambda v: _digits_chave(v) if pd.notna(v) and str(v).strip() else ""
    )
    pref = ["LADO", "CHV_NFE", "NUM_DOC", "SER", "CNPJ", "COD_PART", "VL_DOC", "COD_MOD"]
    outros = [c for c in out.columns if c not in pref]
    ordered = [c for c in pref if c in out.columns] + outros
    out = out[ordered]
    ren = {k: v for k, v in _COL_C100_PLAN.items() if k in out.columns}
    return out.rename(columns=ren)


def _comparativo_tabela3_para_excel(comp: pd.DataFrame) -> pd.DataFrame:
    if comp.empty:
        return comp
    m = {
        "CFOP SPED": "CFOP",
        "Presenca_CFOP": "Presença do CFOP (Cliente / Nascel / só um lado)",
        "Divergente": "Divergente (SIM ou NÃO)",
        "Observação": "Observação (preencher à mão)",
    }
    for col in _COLS_C190_SOMA:
        ant = f"Dif. {col} (Cli−Nascel)"
        leg = _COL_VALOR_PT.get(col, col)
        m[ant] = f"Diferença — {leg} (Cliente − Nascel)"
    return comp.rename(columns=m)


def _onde_agir_cfop_para_excel(onde: pd.DataFrame) -> pd.DataFrame:
    if onde.empty:
        return onde
    m = {
        "#": "N.º",
        "Prioridade": "Prioridade",
        "Chave_44_NF": _LABEL_CHAVE_ONDE_AGIR,
        "NUM_DOC": "Número do documento",
        "CFOP": "CFOP",
        "CST_ICMS": "CST ICMS",
        "Coluna_no_comparativo": "Coluna correspondente no comparativo",
        "Valor_SPED_Cliente": "Valor no SPED Cliente",
        "Valor_SPED_Nascel": "Valor no SPED Nascel",
        "Dif_Cliente_menos_Nascel": "Diferença (Cliente − Nascel)",
        "Acao": "O que fazer / onde conferir",
        "Tipo": "Tipo",
        "Referencia": "Referência",
    }
    cols = {c: m.get(c, c) for c in onde.columns}
    return onde.rename(columns=cols)


# Identidade visual Detetive no Excel — **pink + laranja** (sem verde/azul do tema Office).
_FMT_XLSX_TITULO_BLOCO_PINK = {
    "bold": True,
    "bg_color": "#FBCFE8",
    "font_color": "#831843",
    "font_size": 11,
}
_FMT_XLSX_TITULO_BLOCO_LARANJA = {
    "bold": True,
    "bg_color": "#FED7AA",
    "font_color": "#9A3412",
    "font_size": 11,
}
# Compat: somatório único e títulos que não alternam cor.
_FMT_XLSX_TITULO_BLOCO = _FMT_XLSX_TITULO_BLOCO_PINK

# Light 1 = branco / cinza muito claro — não segue o «accent» do tema (evita verde/azul aleatório).
_XLSX_TABLE_STYLE_DINAMICA = "Table Style Light 1"

_ABA_PINK_HEX = "#DB2777"
_ABA_LARANJA_HEX = "#EA580C"


def _xlsx_aplicar_abas_rosa(writer: pd.ExcelWriter) -> None:
    """Separadores das folhas alternando pink e laranja."""
    for i, (_nm, ws) in enumerate(writer.sheets.items()):
        try:
            ws.set_tab_color(_ABA_PINK_HEX if i % 2 == 0 else _ABA_LARANJA_HEX)
        except Exception:
            pass


# Ordem das abas no Excel comparativo: comparativo → onde agir → detalhe NF-e → bloco CT-e (D).
_ORDEM_ABAS_COMP_C190 = [
    "Comparativo_CFOP_C190",
    "ONDE_AGIR",
    "Cliente",
    "Nascel",
]
_ORDEM_ABAS_COMP_D190 = [
    "Comparativo_CFOP_D190",
    "ONDE_AGIR",
    "D100_Cliente",
    "D100_Nascel",
    "D190_Cliente",
    "D190_Nascel",
]
_ORDEM_ABAS_COMP_NFE_CTE = [
    "Comparativo_CFOP_C190",
    "Comparativo_CFOP_D190",
    "ONDE_AGIR_NF_e",
    "ONDE_AGIR_CTe",
    "Cliente",
    "Nascel",
    "D100_Cliente",
    "D100_Nascel",
    "D190_Cliente",
    "D190_Nascel",
]


def _xlsx_reordenar_abas_por_lista(buf: io.BytesIO, ordem_desejada: list[str]) -> io.BytesIO:
    """Reordena folhas no .xlsx já gravado (openpyxl), mantendo o resto no fim."""
    from openpyxl import load_workbook
    from openpyxl.styles.colors import Color

    buf.seek(0)
    wb = load_workbook(buf)
    presentes = list(wb.sheetnames)
    nova = [n for n in ordem_desejada if n in presentes]
    for n in presentes:
        if n not in nova:
            nova.append(n)
    por_titulo = {ws.title: ws for ws in wb.worksheets}
    wb._sheets = [por_titulo[t] for t in nova]
    # Tabs pink/laranja na ordem **final** (após o xlsxwriter ter gravado noutra ordem).
    pink, lar = "FFDB2777", "FFEA580C"
    for i, ws in enumerate(wb.worksheets):
        try:
            ws.sheet_properties.tabColor = Color(rgb=pink if i % 2 == 0 else lar)
        except Exception:
            pass
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _xlsx_add_tabela_estilo_dinamica(
    ws,
    first_row: int,
    last_row: int,
    first_col: int,
    last_col: int,
    headers: list[str] | tuple[str, ...] | None = None,
) -> None:
    """
    Envolve o intervalo numa tabela com filtro. Sem «columns» com header, o Excel
    mostra «Column1», «Column2»… por cima dos dados — daí os rótulos explícitos.
    """
    if last_row < first_row or last_col < first_col:
        return
    ncols = last_col - first_col + 1
    opts: dict = {
        "style": _XLSX_TABLE_STYLE_DINAMICA,
        "autofilter": True,
        "name": f"T_{uuid.uuid4().hex[:12]}",
    }
    if headers is not None:
        hlist = [
            str(headers[i]) if i < len(headers) else f"Coluna_{i + 1}"
            for i in range(ncols)
        ]
        opts["columns"] = [{"header": h} for h in hlist]
    ws.add_table(first_row, first_col, last_row, last_col, opts)


def _xlsx_tabela_sobre_df_escrito(ws, df: pd.DataFrame) -> None:
    """DataFrame já escrito com header na linha 0: cabeçalho + len(df) linhas de dados."""
    if df is None or len(df.columns) < 1:
        return
    n_linhas_dados = len(df)
    if n_linhas_dados < 0:
        return
    n_cols = len(df.columns)
    last_r = n_linhas_dados  # header row 0 + dados 1..n_linhas_dados
    _xlsx_add_tabela_estilo_dinamica(
        ws, 0, last_r, 0, n_cols - 1, list(df.columns)
    )


def _norm_cfop_sped(v) -> str:
    t = _norm_val(v).replace(".0", "").strip()
    if not t or t.lower() in ("nan", "none"):
        return "(vazio)"
    return t


def _carregar_blocos_txt_via_spedlib(file_obj, filename: str) -> dict[str, pd.DataFrame]:
    """
    Leitura .txt: remove assinatura (|9999|) + EFDReader (spedlib) + reforço posicional
    de VL_DOC / totais (split | como no Guia), para não deslocar valores quando o arquivo
    tem quantidade de campos diferente do layout fixo do reader.
    """
    import os
    import tempfile

    from spedlib.efd_reader import EFDReader
    from spedlib.utils import remove_efd_signature

    file_obj.seek(0)
    raw = file_obj.read()
    if isinstance(raw, str):
        raw = raw.encode("latin-1", errors="replace")

    path_in = None
    path_clean = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".efd.txt") as tmp_in:
            path_in = tmp_in.name
            tmp_in.write(raw)
        path_clean = path_in + ".sem_assinatura.txt"
        remove_efd_signature(path_in, path_clean, encoding="latin-1")
        reader = EFDReader(encoding="latin-1")
        reader.read_file(path_clean)
        d = reader.data
        blocos = {
            "C100": d["C100"].copy(),
            "C190": d["C190"].copy(),
            "D100": d["D100"].copy(),
            "D190": d["D190"].copy(),
        }
        try:
            with open(path_clean, encoding="latin-1") as tf:
                _reforcar_valores_sped_txt_por_posicao(tf.read(), blocos)
        except OSError:
            pass
        return blocos
    finally:
        for p in (path_in, path_clean):
            if p and os.path.isfile(p):
                try:
                    os.unlink(p)
                except OSError:
                    pass
        file_obj.seek(0)


def carregar_blocos_sped_completos(file_obj, filename: str) -> dict[str, pd.DataFrame]:
    """C100, C190, D100, D190 — .txt via spedlib (igual ao teu exportador); .xlsx via abas."""
    import confronto_gerencial as cg

    fn = (filename or "").lower()
    file_obj.seek(0)
    if fn.endswith(".txt"):
        return _carregar_blocos_txt_via_spedlib(file_obj, filename)
    mapa = cg.carregar_mapa_sped(file_obj, filename)
    file_obj.seek(0)
    c100, c190 = cg._df_c100_c190_de_mapa_sped(mapa)
    d100, d190 = cg._df_d100_d190_de_mapa_sped(mapa)
    return {
        "C100": c100.copy(),
        "C190": c190.copy(),
        "D100": d100.copy(),
        "D190": d190.copy(),
    }


def carregar_df_c190_de_arquivo(file_obj, filename: str) -> pd.DataFrame:
    """C190 já ligado ao C100 (hierarquia .txt ou merge no Excel)."""
    bl = carregar_blocos_sped_completos(file_obj, filename)
    return bl["C190"].copy()


def _agg_cfop_sem_total(df_c190: pd.DataFrame) -> pd.DataFrame:
    if df_c190.empty or "CFOP" not in df_c190.columns:
        return pd.DataFrame(columns=["CFOP SPED"] + _COLS_C190_SOMA)
    g = df_c190.copy()
    for c in _COLS_C190_SOMA:
        if c not in g.columns:
            g[c] = 0.0
        else:
            g[c] = pd.to_numeric(g[c], errors="coerce").fillna(0.0)
    g["CFOP SPED"] = g["CFOP"].map(_norm_cfop_sped)
    agg = g.groupby("CFOP SPED", dropna=False)[_COLS_C190_SOMA].sum().reset_index()
    return agg.round(2)


def tabela_somatorio_cfop_estilo_excel(df_c190: pd.DataFrame) -> pd.DataFrame:
    """Tabela dinâmica por CFOP com cabeçalhos em português (CFOP + Soma — …)."""
    agg = _agg_cfop_sem_total(df_c190)
    if agg.empty:
        return pd.DataFrame(columns=_CAB_CFOP_EXCEL)
    ren: dict[str, str] = {"CFOP SPED": "CFOP"}
    for k in _COLS_C190_SOMA:
        if k in agg.columns:
            ren[k] = _SOMA_COL_PT[k]
    out = agg.rename(columns=ren)
    tot = out.drop(columns=["CFOP"]).sum(numeric_only=True)
    total_row: dict = {"CFOP": "Total Geral"}
    for c in out.columns:
        if c != "CFOP":
            total_row[c] = round(float(tot.get(c, 0.0)), 2)
    out = pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)
    return out


def _comparativo_cfop_merge(agg_c: pd.DataFrame, agg_n: pd.DataFrame) -> pd.DataFrame:
    if agg_c.empty and agg_n.empty:
        return pd.DataFrame()
    set_c = set(agg_c["CFOP SPED"]) if not agg_c.empty else set()
    set_n = set(agg_n["CFOP SPED"]) if not agg_n.empty else set()
    m = agg_c.merge(agg_n, on="CFOP SPED", how="outer", suffixes=("_Cliente", "_Nascel")).fillna(0.0)
    linhas = []
    tol = 0.02
    for _, row in m.iterrows():
        cf = row["CFOP SPED"]
        if cf in set_c and cf in set_n:
            pres = "Ambos"
        elif cf in set_c:
            pres = f"Só {ROTULO_CLIENTE}"
        else:
            pres = f"Só {ROTULO_NASCEL}"
        r = {"CFOP SPED": cf, "Presenca_CFOP": pres}
        div = pres != "Ambos"
        for col in _COLS_C190_SOMA:
            vc = float(row.get(f"{col}_Cliente", 0.0))
            vn = float(row.get(f"{col}_Nascel", 0.0))
            d = round(vc - vn, 2)
            # Só diferença (Cliente − Nascel); totais estão nas Tabelas 1 e 2
            r[f"Dif. {col} (Cli−Nascel)"] = d
            if abs(d) >= tol:
                div = True
        r["Divergente"] = "SIM" if div else "NÃO"
        r["Observação"] = ""
        linhas.append(r)
    out = pd.DataFrame(linhas)
    if not out.empty:
        out = out.sort_values("CFOP SPED")
    return out


def _sheet_ref_excel(nome: str) -> str:
    n = (nome or "")[:31].replace("'", "''")
    return f"'{n}'"


def _refs_sumifs(
    sheet: str,
    col_sum_letter: str,
    col_crit_letter: str,
    excel_row_first: int,
    excel_row_last: int,
) -> tuple[str, str]:
    esc = _sheet_ref_excel(sheet)
    s = f"{esc}!${col_sum_letter}${excel_row_first}:${col_sum_letter}${excel_row_last}"
    c = f"{esc}!${col_crit_letter}${excel_row_first}:${col_crit_letter}${excel_row_last}"
    return s, c


def _pode_formulas_somatorio(
    df: pd.DataFrame,
    mapa: dict[str, str] | None = None,
) -> bool:
    if df is None or df.empty:
        return False
    m = mapa or _COL_DETALHE_C190
    if _col_excel_idx(df, _COL_CFOP_SOM, m) is None:
        return False
    return all(_col_excel_idx(df, c, m) is not None for c in _COLS_C190_SOMA)


def _letras_colunas_soma(
    df: pd.DataFrame,
    mapa: dict[str, str] | None = None,
) -> dict[str, str] | None:
    m = mapa or _COL_DETALHE_C190
    if not _pode_formulas_somatorio(df, m):
        return None
    out: dict[str, str] = {}
    for c in _COLS_C190_SOMA:
        j = _col_excel_idx(df, c, m)
        if j is None:
            return None
        out[c] = xl_col_to_name(j)
    jc = _col_excel_idx(df, _COL_CFOP_SOM, m)
    if jc is None:
        return None
    out["_crit"] = xl_col_to_name(jc)
    return out


def _xlsx_reforcar_detalhe_texto_e_numero(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
    mapa: dict[str, str] | None = None,
) -> None:
    """Coluna de CFOP texto @; colunas de valor como número (SUMIFS)."""
    m = mapa or _COL_DETALHE_C190
    if df is None or df.empty:
        return
    ix_cf = _col_excel_idx(df, _COL_CFOP_SOM, m)
    if ix_cf is None:
        return
    ws = writer.sheets[sheet_name]
    wb = writer.book
    fmt_txt = wb.add_format({"num_format": "@"})
    fmt_num = wb.add_format({"num_format": "#,##0.00"})
    n = len(df)
    for r in range(n):
        ws.write_string(r + 1, ix_cf, str(df.iloc[r, ix_cf]), fmt_txt)
    for cname in _COLS_C190_SOMA:
        j = _col_excel_idx(df, cname, m)
        if j is None:
            continue
        for r in range(n):
            raw = df.iloc[r, j]
            v = pd.to_numeric(raw, errors="coerce")
            num = 0.0 if pd.isna(v) else float(v)
            ws.write_number(r + 1, j, num, fmt_num)


def _xlsx_reforcar_cfop_sped_coluna_a_texto(
    ws,
    tab: pd.DataFrame,
    start_header_row: int,
    fmt_txt,
) -> None:
    """Coluna A em formato texto @; CFOP igual ao detalhe (norm) exceto linha Total Geral."""
    if tab.empty:
        return
    n = len(tab)
    col_cfop = "CFOP" if "CFOP" in tab.columns else "CFOP SPED"
    for i in range(n):
        xrow = start_header_row + 1 + i
        cf = tab.iloc[i][col_cfop]
        if str(cf).strip() == "Total Geral":
            txt = "Total Geral"
        else:
            txt = str(_norm_cfop_sped(cf))
        ws.write_string(xrow, 0, txt, fmt_txt)


def _sobrescrever_somatorio_com_formulas(
    ws,
    tab: pd.DataFrame,
    start_header_row: int,
    sheet_dados: str,
    letras: dict[str, str],
    excel_row_dados_ini: int,
    excel_row_dados_fim: int,
    fmt_num,
) -> None:
    if tab.empty or len(tab) < 2:
        return
    hdr = list(tab.columns)
    n = len(tab)
    for i in range(n - 1):
        xrow = start_header_row + 1 + i
        crit_cell = xl_rowcol_to_cell(xrow, 0, row_abs=True, col_abs=True)
        row_vals = tab.iloc[i]
        for py_col in _COLS_C190_SOMA:
            col_excel = _SOMA_COL_PT[py_col]
            if col_excel not in hdr:
                continue
            j = hdr.index(col_excel)
            sum_rng, crit_rng = _refs_sumifs(
                sheet_dados,
                letras[py_col],
                letras["_crit"],
                excel_row_dados_ini,
                excel_row_dados_fim,
            )
            fmla = f"=SUMIFS({sum_rng},{crit_rng},{crit_cell})"
            try:
                cached = float(row_vals[col_excel])
            except (TypeError, ValueError):
                cached = 0.0
            ws.write_formula(xrow, j, fmla, fmt_num, cached)
    xrow_tot = start_header_row + 1 + (n - 1)
    r_first = start_header_row + 2
    r_last = start_header_row + n
    tot_vals = tab.iloc[n - 1]
    for py_col in _COLS_C190_SOMA:
        col_excel = _SOMA_COL_PT[py_col]
        if col_excel not in hdr:
            continue
        j = hdr.index(col_excel)
        cl = xl_col_to_name(j)
        # 9 = SOMA; ignora linhas ocultas pelo filtro automático nesta folha
        ftot = f"=SUBTOTAL(9,{cl}{r_first}:{cl}{r_last})"
        try:
            cached_t = float(tot_vals[col_excel])
        except (TypeError, ValueError):
            cached_t = 0.0
        ws.write_formula(xrow_tot, j, ftot, fmt_num, cached_t)


def _escrever_comparativo_cfop_tres_tabelas(
    writer: pd.ExcelWriter,
    tab_cli: pd.DataFrame,
    tab_nas: pd.DataFrame,
    tab_diff: pd.DataFrame,
    nome_sheet: str = "Comparativo_CFOP",
    bloco_fonte: str = "C190",
    *,
    sheet_dados_cliente: str = "",
    sheet_dados_nascel: str = "",
    df_cli_export: pd.DataFrame | None = None,
    df_nas_export: pd.DataFrame | None = None,
) -> None:
    """
    Três blocos: T1/T2 com SUMIFS nas abas de detalhe após forçar CFOP como texto e VL_* como número.
    Tabela 3 continua em valores (Python).
    """
    sh = nome_sheet[:31]
    fmt_pink = writer.book.add_format(_FMT_XLSX_TITULO_BLOCO_PINK)
    fmt_laranja = writer.book.add_format(_FMT_XLSX_TITULO_BLOCO_LARANJA)
    fmt_txt_cfop = writer.book.add_format({"num_format": "@"})
    fmt_num_soma = writer.book.add_format({"num_format": "#,##0.00"})
    ncol1 = max(len(tab_cli.columns), 1)
    ncol2 = max(len(tab_nas.columns), 1)
    tab_diff_x = _comparativo_tabela3_para_excel(tab_diff)
    ncol3 = max(len(tab_diff_x.columns), 1)

    mapa_det = _COL_DETALHE_D190 if bloco_fonte == "D190" else _COL_DETALHE_C190

    start1 = 1
    tab_cli.to_excel(writer, sheet_name=sh, index=False, startrow=start1)
    ws = writer.sheets[sh]
    ws.merge_range(
        0,
        0,
        0,
        ncol1 - 1,
        f"Tabela 1 — Somatório por CFOP ({bloco_fonte}) — {ROTULO_CLIENTE}",
        fmt_pink,
    )
    _xlsx_reforcar_cfop_sped_coluna_a_texto(ws, tab_cli, start1, fmt_txt_cfop)

    lc = (
        _letras_colunas_soma(df_cli_export, mapa_det)
        if df_cli_export is not None
        else None
    )
    ln = (
        _letras_colunas_soma(df_nas_export, mapa_det)
        if df_nas_export is not None
        else None
    )
    usar_f = (
        bool(sheet_dados_cliente and sheet_dados_nascel and lc and ln)
        and df_cli_export is not None
        and df_nas_export is not None
        and len(df_cli_export) > 0
        and len(df_nas_export) > 0
    )
    if usar_f:
        ex_ini = 2
        ex_fim_cli = 1 + len(df_cli_export)
        ex_fim_nas = 1 + len(df_nas_export)
        _sobrescrever_somatorio_com_formulas(
            ws,
            tab_cli,
            start1,
            sheet_dados_cliente,
            lc,
            ex_ini,
            ex_fim_cli,
            fmt_num_soma,
        )

    if len(tab_cli) > 0:
        _xlsx_add_tabela_estilo_dinamica(
            ws,
            start1,
            start1 + len(tab_cli),
            0,
            ncol1 - 1,
            list(tab_cli.columns),
        )

    start2 = start1 + len(tab_cli) + 3
    tab_nas.to_excel(writer, sheet_name=sh, index=False, startrow=start2)
    ws.merge_range(
        start2 - 1,
        0,
        start2 - 1,
        ncol2 - 1,
        f"Tabela 2 — Somatório por CFOP ({bloco_fonte}) — {ROTULO_NASCEL}",
        fmt_laranja,
    )
    _xlsx_reforcar_cfop_sped_coluna_a_texto(ws, tab_nas, start2, fmt_txt_cfop)

    if usar_f:
        _sobrescrever_somatorio_com_formulas(
            ws,
            tab_nas,
            start2,
            sheet_dados_nascel,
            ln,
            ex_ini,
            ex_fim_nas,
            fmt_num_soma,
        )

    if len(tab_nas) > 0:
        _xlsx_add_tabela_estilo_dinamica(
            ws,
            start2,
            start2 + len(tab_nas),
            0,
            ncol2 - 1,
            list(tab_nas.columns),
        )

    start3 = start2 + len(tab_nas) + 3
    tab_diff_x.to_excel(writer, sheet_name=sh, index=False, startrow=start3)
    ws.merge_range(
        start3 - 1,
        0,
        start3 - 1,
        ncol3 - 1,
        f"Tabela 3 ({bloco_fonte}) — Diferenças (Cliente − Nascel), presença, divergência e observação",
        fmt_pink,
    )
    ws.set_column(0, min(25, ncol3), 16)
    obs_tit = "Observação (preencher à mão)"
    if obs_tit in tab_diff_x.columns:
        j_obs = list(tab_diff_x.columns).index(obs_tit)
        ws.set_column(j_obs, j_obs, 52)
    if len(tab_diff_x) > 0:
        _xlsx_add_tabela_estilo_dinamica(
            ws,
            start3,
            start3 + len(tab_diff_x),
            0,
            ncol3 - 1,
            list(tab_diff_x.columns),
        )


def _escrever_somatorio_cfop_unica(
    writer: pd.ExcelWriter,
    tab: pd.DataFrame,
    nome_sheet: str,
    bloco_fonte: str,
    rotulo_lado: str,
    sheet_dados: str,
    df_export: pd.DataFrame | None,
) -> None:
    """Uma só tabela tipo «Tabela 1» do comparativo, com SUMIFS no detalhe quando houver linhas."""
    sh = nome_sheet[:31]
    fmt = writer.book.add_format(_FMT_XLSX_TITULO_BLOCO)
    fmt_txt_cfop = writer.book.add_format({"num_format": "@"})
    fmt_num_soma = writer.book.add_format({"num_format": "#,##0.00"})
    ncol = max(len(tab.columns), 1)
    start = 1
    tab.to_excel(writer, sheet_name=sh, index=False, startrow=start)
    ws = writer.sheets[sh]
    ws.merge_range(
        0,
        0,
        0,
        ncol - 1,
        f"Tabela 1 — Somatório por CFOP ({bloco_fonte}) — {rotulo_lado}",
        fmt,
    )
    _xlsx_reforcar_cfop_sped_coluna_a_texto(ws, tab, start, fmt_txt_cfop)
    mapa_det = _COL_DETALHE_D190 if bloco_fonte == "D190" else _COL_DETALHE_C190
    letras = (
        _letras_colunas_soma(df_export, mapa_det) if df_export is not None else None
    )
    if (
        sheet_dados
        and letras
        and df_export is not None
        and len(df_export) > 0
    ):
        ex_ini = 2
        ex_fim = 1 + len(df_export)
        _sobrescrever_somatorio_com_formulas(
            ws,
            tab,
            start,
            sheet_dados,
            letras,
            ex_ini,
            ex_fim,
            fmt_num_soma,
        )
    if len(tab) > 0:
        _xlsx_add_tabela_estilo_dinamica(
            ws, start, start + len(tab), 0, ncol - 1, list(tab.columns)
        )
    ws.set_column(0, min(25, ncol), 16)


def _digits_chave(v) -> str:
    return re.sub(r"\D", "", _norm_val(v))


def _garantir_chv_em_c190(c100: pd.DataFrame, c190: pd.DataFrame) -> pd.DataFrame:
    """SPED .txt (spedlib): C190 não traz CHV_NFE nas colunas — cruza com C100 por NUM/SER/PART."""
    if c190.empty:
        return c190
    if "CHV_NFE" in c190.columns:
        s = c190["CHV_NFE"].map(_digits_chave)
        if s.astype(str).str.len().ge(40).any():
            return c190
    keys = [k for k in ("NUM_DOC", "SER", "COD_PART") if k in c100.columns and k in c190.columns]
    if len(keys) < 2 or c100.empty or "CHV_NFE" not in c100.columns:
        out = c190.copy()
        if "CHV_NFE" not in out.columns:
            out["CHV_NFE"] = ""
        return out
    lkp = c100[keys + ["CHV_NFE"]].copy()
    lkp["CHV_NFE"] = lkp["CHV_NFE"].map(_digits_chave)
    lkp = lkp.drop_duplicates(subset=keys, keep="first")
    base = c190.drop(columns=["CHV_NFE"], errors="ignore")
    return base.merge(lkp, on=keys, how="left")


def _garantir_chv_em_d190(d100: pd.DataFrame, d190: pd.DataFrame) -> pd.DataFrame:
    if d190.empty:
        return d190
    for c in ("CHV_DOC", "CHV_NFE"):
        if c in d190.columns:
            s = d190[c].map(_digits_chave)
            if s.astype(str).str.len().ge(40).any():
                return d190
            break
    keys = [k for k in ("NUM_DOC", "SER", "COD_PART") if k in d100.columns and k in d190.columns]
    ch_col = "CHV_DOC" if "CHV_DOC" in d100.columns else ("CHV_NFE" if "CHV_NFE" in d100.columns else None)
    if len(keys) < 2 or d100.empty or not ch_col:
        return d190
    lkp = d100[keys + [ch_col]].copy()
    lkp[ch_col] = lkp[ch_col].map(_digits_chave)
    lkp = lkp.rename(columns={ch_col: "CHV_DOC"})
    lkp = lkp.drop_duplicates(subset=keys, keep="first")
    base = d190.drop(columns=[c for c in ("CHV_DOC", "CHV_NFE") if c in d190.columns], errors="ignore")
    out = base.merge(lkp, on=keys, how="left")
    return out


def _serie_chv44(df: pd.DataFrame, fonte_cfop: str | None = None) -> pd.Series:
    for c in ("CHV_DOC", "CHV_NFE", "CHV_CTE"):
        if c in df.columns:
            return df[c].map(_digits_chave)
    return pd.Series([""] * len(df), index=df.index, dtype=object)


def _agregar_por_chave_cfop_cst(df: pd.DataFrame, fonte_cfop: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["CHV_44", "NUM_DOC", "CFOP SPED", "CST_ICMS"] + _COLS_C190_SOMA)
    g = df.copy()
    g["CHV_44"] = _serie_chv44(g, fonte_cfop)
    if "NUM_DOC" not in g.columns:
        g["NUM_DOC"] = ""
    g["NUM_DOC"] = g["NUM_DOC"].map(_norm_val)
    g["CFOP SPED"] = g["CFOP"].map(_norm_cfop_sped) if "CFOP" in g.columns else "(vazio)"
    g["CST_ICMS"] = g["CST_ICMS"].map(_norm_val) if "CST_ICMS" in g.columns else ""
    for c in _COLS_C190_SOMA:
        if c not in g.columns:
            g[c] = 0.0
        else:
            g[c] = pd.to_numeric(g[c], errors="coerce").fillna(0.0)
    agg = {c: "sum" for c in _COLS_C190_SOMA}
    agg["NUM_DOC"] = "first"
    return g.groupby(["CHV_44", "CFOP SPED", "CST_ICMS"], dropna=False).agg(agg).reset_index()


def _montar_onde_agir_cfop_por_nota(
    df_cli: pd.DataFrame,
    df_nas: pd.DataFrame,
    fonte_cfop: str,
    max_linhas: int = 500_000,
) -> pd.DataFrame:
    """
    Uma linha por combinação (chave 44, CFOP, CST, coluna de valor) com diferença ≥ 0,02.
    Indica a coluna homóloga na Tabela 3 do comparativo.
    """
    pc = _agregar_por_chave_cfop_cst(df_cli, fonte_cfop)
    pn = _agregar_por_chave_cfop_cst(df_nas, fonte_cfop)
    if pc.empty and pn.empty:
        return pd.DataFrame()

    keys = ["CHV_44", "CFOP SPED", "CST_ICMS"]
    m = pc.merge(pn, on=keys, how="outer", suffixes=("_Cliente", "_Nascel"))
    for col in _COLS_C190_SOMA:
        cl, cn = f"{col}_Cliente", f"{col}_Nascel"
        if cl not in m.columns:
            m[cl] = 0.0
        if cn not in m.columns:
            m[cn] = 0.0
        m[cl] = pd.to_numeric(m[cl], errors="coerce").fillna(0.0)
        m[cn] = pd.to_numeric(m[cn], errors="coerce").fillna(0.0)

    ndc = "NUM_DOC_Cliente" if "NUM_DOC_Cliente" in m.columns else None
    ndn = "NUM_DOC_Nascel" if "NUM_DOC_Nascel" in m.columns else None
    tol = 0.02
    det_aba = (
        "«Cliente» e «Nascel»"
        if fonte_cfop == "C190"
        else "«D190_Cliente» e «D190_Nascel»"
    )
    linhas: list[dict] = []
    truncou = False
    for _, row in m.iterrows():
        chv = str(row["CHV_44"]).strip()
        cfop = row["CFOP SPED"]
        cst = row.get("CST_ICMS", "")
        num_doc = ""
        if ndc and pd.notna(row.get(ndc)):
            num_doc = str(row[ndc]).strip()
        if not num_doc and ndn and pd.notna(row.get(ndn)):
            num_doc = str(row[ndn]).strip()

        for col in _COLS_C190_SOMA:
            vc = float(row[f"{col}_Cliente"])
            vn = float(row[f"{col}_Nascel"])
            d = round(vc - vn, 2)
            if abs(d) < tol:
                continue
            if len(linhas) >= max_linhas:
                truncou = True
                break
            col_comp = f"Tabela 3 — Dif. {col} (Cli−Nascel)"
            prio = "Alta" if col in ("VL_BC_ICMS", "VL_ICMS", "VL_OPR") else "Média"
            ch_exibe = chv if len(chv) >= 30 else (chv if chv else "— (use NUM_DOC ou abra o SPED .txt com C100)")
            linhas.append(
                {
                    "#": len(linhas) + 1,
                    "Prioridade": prio,
                    "Chave_44_NF": ch_exibe,
                    "NUM_DOC": num_doc,
                    "CFOP": cfop,
                    "CST_ICMS": cst,
                    "Coluna_no_comparativo": col_comp,
                    "Valor_SPED_Cliente": round(vc, 2),
                    "Valor_SPED_Nascel": round(vn, 2),
                    "Dif_Cliente_menos_Nascel": d,
                    "Acao": (
                        f"Localizar esta NF/chave nas abas {det_aba} "
                        f"(filtrar pela chave ou pelo número do documento). A coluna acima é a mesma métrica que em "
                        f"«{col_comp}» na folha de comparativo (por CFOP agregado vs aqui por documento+CFOP+CST)."
                    ),
                }
            )
        if truncou:
            break

    if truncou:
        linhas.append(
            {
                "#": len(linhas) + 1,
                "Prioridade": "—",
                "Chave_44_NF": "—",
                "NUM_DOC": "",
                "CFOP": "",
                "CST_ICMS": "",
                "Coluna_no_comparativo": "—",
                "Valor_SPED_Cliente": "",
                "Valor_SPED_Nascel": "",
                "Dif_Cliente_menos_Nascel": "",
                "Acao": f"Lista limitada a {max_linhas} linhas; há mais divergências — filtrar no SPED ou reduzir período.",
            }
        )
    return pd.DataFrame(linhas)


def _montar_onde_agir_cfop_somente_agregado(comp: pd.DataFrame, bloco_fonte: str) -> pd.DataFrame:
    div = comp[comp["Divergente"] == "SIM"]
    if div.empty:
        return pd.DataFrame(
            [
                {
                    "#": 1,
                    "Prioridade": "—",
                    "Tipo": "Somatório CFOP",
                    "Referencia": "Todos os CFOP",
                    "Acao": "Nenhuma diferença acima de R$ 0,02 entre os dois SPED (por CFOP).",
                }
            ]
        )
    linhas = []
    for i, (_, row) in enumerate(div.iterrows(), start=1):
        cf = row["CFOP SPED"]
        pres = row.get("Presenca_CFOP", "Ambos")
        partes = []
        for col in _COLS_C190_SOMA:
            d = row.get(f"Dif. {col} (Cli−Nascel)", 0)
            if abs(float(d)) >= 0.02:
                partes.append(f"{col} dif={d}")
        if pres != "Ambos":
            acao = f"{pres}. Conferir se o CFOP existe só num dos SPED ou foi mapeado errado."
            if partes:
                acao += " Valores: " + "; ".join(partes)
        else:
            acao = "Conferir notas com este CFOP nos dois SPED: " + ("; ".join(partes) if partes else "(ver Tabela 3)")
        linhas.append(
            {
                "#": i,
                "Prioridade": "Alta",
                "Tipo": "CFOP com somatório diferente" if pres == "Ambos" else "CFOP só num dos SPED",
                "Referencia": str(cf),
                "Acao": acao,
            }
        )
    return pd.DataFrame(linhas)


def montar_onde_agir_cfop(
    comp: pd.DataFrame,
    bloco_fonte: str = "C190",
    df_cli: pd.DataFrame | None = None,
    df_nas: pd.DataFrame | None = None,
) -> pd.DataFrame:
    if comp.empty:
        hint = (
            "Sem linhas D190 — use .txt EFD com D100 seguido de D190, ou Excel com abas D100 e D190."
            if bloco_fonte == "D190"
            else "Sem linhas C190 — use .txt EFD com C100 seguido de C190, ou Excel com abas C100 e C190."
        )
        return pd.DataFrame(
            [
                {
                    "#": 1,
                    "Prioridade": "—",
                    "Tipo": f"CFOP ({bloco_fonte})",
                    "Referencia": "—",
                    "Acao": hint,
                }
            ]
        )

    if df_cli is not None and df_nas is not None and (not df_cli.empty or not df_nas.empty):
        det = _montar_onde_agir_cfop_por_nota(df_cli, df_nas, bloco_fonte)
        if not det.empty:
            return det

    return _montar_onde_agir_cfop_somente_agregado(comp, bloco_fonte)


def _df_com_coluna_lado(df: pd.DataFrame, lado: str) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame(columns=["LADO"])
    if df.empty:
        return pd.DataFrame(columns=["LADO"] + list(df.columns))
    out = df.copy()
    out.insert(0, "LADO", lado)
    return out


def _df_export_sped_com_cfop_soma(df: pd.DataFrame, lado: str) -> pd.DataFrame:
    """Inclui CFOP_SPED_soma (texto) para o critério do SUMIFS coincidir com a coluna A do comparativo."""
    x = _df_com_coluna_lado(df, lado).copy()
    if x.empty:
        x[_COL_CFOP_SOM] = pd.Series(dtype=object)
        return x
    if "CFOP" in x.columns:
        x[_COL_CFOP_SOM] = x["CFOP"].map(_norm_cfop_sped).astype(str)
    else:
        x[_COL_CFOP_SOM] = "(vazio)"
    return x


def _ler_bytes_upload_sped(file_obj) -> bytes:
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)
    return file_obj.read()


def _carregar_blocos_de_bytes(data: bytes, filename: str) -> dict[str, pd.DataFrame]:
    return carregar_blocos_sped_completos(io.BytesIO(data), filename)


def _carregar_dois_speds_em_paralelo(
    file_sped_cliente,
    file_sped_nascel,
    nome_arquivo_cliente: str,
    nome_arquivo_nascel: str,
) -> tuple[dict[str, pd.DataFrame], dict[str, pd.DataFrame]]:
    """
    Copia cada upload para memória e faz o parse dos dois SPED em paralelo (dois ficheiros ao mesmo tempo).
    Isto é só desempenho: cada SPED continua independente — não funde nem soma os dois num único conjunto de dados.
    """
    dc = _ler_bytes_upload_sped(file_sped_cliente)
    dn = _ler_bytes_upload_sped(file_sped_nascel)
    with ThreadPoolExecutor(max_workers=2) as ex:
        fut_c = ex.submit(_carregar_blocos_de_bytes, dc, nome_arquivo_cliente)
        fut_n = ex.submit(_carregar_blocos_de_bytes, dn, nome_arquivo_nascel)
        bl_c = fut_c.result()
        bl_n = fut_n.result()
    return bl_c, bl_n


def _escrever_um_bloco_comparativo_cfop(
    writer: pd.ExcelWriter,
    bl_c: dict[str, pd.DataFrame],
    bl_n: dict[str, pd.DataFrame],
    fonte_cfop: str,
    onde_sheet: str,
) -> bool:
    """
    Escreve no mesmo workbook um bloco completo: detalhe (C190 NF-e ou D190 CT-e),
    folha Comparativo_CFOP_* e folha «onde agir» com nome `onde_sheet`.
    Devolve False se não houver linhas C190/D190 em nenhum dos dois SPED (bloco ignorado).
    """
    if fonte_cfop == "C190":
        bl_c["C190"] = _garantir_chv_em_c190(bl_c["C100"], bl_c["C190"])
        bl_n["C190"] = _garantir_chv_em_c190(bl_n["C100"], bl_n["C190"])
    else:
        bl_c["D190"] = _garantir_chv_em_d190(bl_c["D100"], bl_c["D190"])
        bl_n["D190"] = _garantir_chv_em_d190(bl_n["D100"], bl_n["D190"])

    chave_df = "D190" if fonte_cfop == "D190" else "C190"
    df_c = bl_c[chave_df]
    df_n = bl_n[chave_df]

    if df_c.empty and df_n.empty:
        return False

    tab_c = tabela_somatorio_cfop_estilo_excel(df_c)
    tab_n = tabela_somatorio_cfop_estilo_excel(df_n)
    agg_c = _agg_cfop_sem_total(df_c)
    agg_n = _agg_cfop_sem_total(df_n)
    comp = _comparativo_cfop_merge(agg_c, agg_n)
    onde = montar_onde_agir_cfop(
        comp,
        bloco_fonte=fonte_cfop,
        df_cli=df_c,
        df_nas=df_n,
    )

    d100_cli = _df_com_coluna_lado(bl_c["D100"], ROTULO_CLIENTE)
    d100_nas = _df_com_coluna_lado(bl_n["D100"], ROTULO_NASCEL)
    d100_cli_pl = _df_c100_planilha(d100_cli)
    d100_nas_pl = _df_c100_planilha(d100_nas)
    c190_cli = _df_export_c190_planilha(bl_c["C190"], ROTULO_CLIENTE)
    c190_nas = _df_export_c190_planilha(bl_n["C190"], ROTULO_NASCEL)
    d190_cli = _df_export_d190_planilha(bl_c["D190"], ROTULO_CLIENTE)
    d190_nas = _df_export_d190_planilha(bl_n["D190"], ROTULO_NASCEL)

    nome_comp = "Comparativo_CFOP_C190" if fonte_cfop == "C190" else "Comparativo_CFOP_D190"

    if fonte_cfop == "C190":
        sh_c190_cli, sh_c190_nas = "Cliente", "Nascel"
        c190_cli.to_excel(writer, index=False, sheet_name=sh_c190_cli)
        c190_nas.to_excel(writer, index=False, sheet_name=sh_c190_nas)
        _xlsx_reforcar_detalhe_texto_e_numero(
            writer, sh_c190_cli, c190_cli, _COL_DETALHE_C190
        )
        _xlsx_reforcar_detalhe_texto_e_numero(
            writer, sh_c190_nas, c190_nas, _COL_DETALHE_C190
        )
        if len(c190_cli) > 0:
            _xlsx_tabela_sobre_df_escrito(writer.sheets[sh_c190_cli], c190_cli)
        if len(c190_nas) > 0:
            _xlsx_tabela_sobre_df_escrito(writer.sheets[sh_c190_nas], c190_nas)
        abas_largura = (sh_c190_cli, sh_c190_nas)
        sheet_dados_c, sheet_dados_n = sh_c190_cli, sh_c190_nas
    else:
        d100_cli_pl.to_excel(writer, index=False, sheet_name="D100_Cliente")
        d100_nas_pl.to_excel(writer, index=False, sheet_name="D100_Nascel")
        d190_cli.to_excel(writer, index=False, sheet_name="D190_Cliente")
        d190_nas.to_excel(writer, index=False, sheet_name="D190_Nascel")
        _xlsx_reforcar_detalhe_texto_e_numero(
            writer, "D190_Cliente", d190_cli, _COL_DETALHE_D190
        )
        _xlsx_reforcar_detalhe_texto_e_numero(
            writer, "D190_Nascel", d190_nas, _COL_DETALHE_D190
        )
        for nm, dff in (
            ("D100_Cliente", d100_cli_pl),
            ("D100_Nascel", d100_nas_pl),
            ("D190_Cliente", d190_cli),
            ("D190_Nascel", d190_nas),
        ):
            if len(dff) > 0:
                _xlsx_tabela_sobre_df_escrito(writer.sheets[nm], dff)
        abas_largura = (
            "D100_Cliente",
            "D100_Nascel",
            "D190_Cliente",
            "D190_Nascel",
        )
        sheet_dados_c, sheet_dados_n = "D190_Cliente", "D190_Nascel"

    _escrever_comparativo_cfop_tres_tabelas(
        writer,
        tab_c,
        tab_n,
        comp,
        nome_sheet=nome_comp,
        bloco_fonte=fonte_cfop,
        sheet_dados_cliente=sheet_dados_c,
        sheet_dados_nascel=sheet_dados_n,
        df_cli_export=c190_cli if fonte_cfop == "C190" else d190_cli,
        df_nas_export=c190_nas if fonte_cfop == "C190" else d190_nas,
    )

    onde_pl = _onde_agir_cfop_para_excel(onde)
    sh_onde = onde_sheet[:31]
    onde_pl.to_excel(writer, index=False, sheet_name=sh_onde)

    for name in abas_largura:
        if name in writer.sheets:
            writer.sheets[name].set_column(0, 40, 14)
    if fonte_cfop == "C190":
        _xlsx_largura_coluna_chave(writer, "Cliente", c190_cli, _COL_DETALHE_C190)
        _xlsx_largura_coluna_chave(writer, "Nascel", c190_nas, _COL_DETALHE_C190)
    else:
        for sn, dff, mp in (
            ("D100_Cliente", d100_cli_pl, _COL_C100_PLAN),
            ("D100_Nascel", d100_nas_pl, _COL_C100_PLAN),
            ("D190_Cliente", d190_cli, _COL_DETALHE_D190),
            ("D190_Nascel", d190_nas, _COL_DETALHE_D190),
        ):
            _xlsx_largura_coluna_chave(writer, sn, dff, mp)
    if sh_onde in writer.sheets:
        w = writer.sheets[sh_onde]
        ch_tit = _LABEL_CHAVE_ONDE_AGIR
        acao_tit = "O que fazer / onde conferir"
        ncols = len(onde_pl.columns)
        for j in range(ncols):
            hdr = str(onde_pl.columns[j])
            if hdr == ch_tit:
                w.set_column(j, j, 48)
            elif hdr == acao_tit:
                w.set_column(j, j, 52)
            else:
                w.set_column(j, j, 14)
    return True


def _gerar_excel_cfop_comparativo_de_blocos(
    bl_c: dict[str, pd.DataFrame],
    bl_n: dict[str, pd.DataFrame],
    fonte_cfop: str,
):
    """
    Gera o Excel de comparativo CFOP a partir de blocos já carregados.
    Ordem final das abas: Comparativo_CFOP_* → ONDE_AGIR → detalhe (Cliente/Nascel ou D100/D190).
    Altera bl_c/bl_n no ramo correspondente (CHV em C190 ou D190).
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        ok = _escrever_um_bloco_comparativo_cfop(
            writer, bl_c, bl_n, fonte_cfop, "ONDE_AGIR"
        )
        if not ok:
            if fonte_cfop == "D190":
                return None, (
                    "Não encontrei registros D190. No .txt, cada D190 deve vir após o D100 (CT-e); "
                    "no Excel, use abas D100 e D190."
                )
            return None, (
                "Não encontrei registros C190. No .txt, o C190 deve vir logo após o C100; "
                "no Excel, use abas C100 e C190 com colunas padrão EFD."
            )
        _xlsx_aplicar_abas_rosa(writer)
    output.seek(0)
    ordem = (
        _ORDEM_ABAS_COMP_D190 if fonte_cfop == "D190" else _ORDEM_ABAS_COMP_C190
    )
    output = _xlsx_reordenar_abas_por_lista(output, ordem)
    return output, "Sucesso"


def _gerar_excel_cfop_comparativo_nfe_cte_de_blocos(
    bl_c: dict[str, pd.DataFrame],
    bl_n: dict[str, pd.DataFrame],
):
    """
    Um único .xlsx: comparativo por CFOP para **NF-e (C190)** e para **CT-e (D190)**.
    Ordem das abas: ambos Comparativo_CFOP → ambos Onde agir → Cliente/Nascel → D100/D190.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        ok_nfe = _escrever_um_bloco_comparativo_cfop(
            writer, bl_c, bl_n, "C190", "ONDE_AGIR_NF_e"
        )
        ok_cte = _escrever_um_bloco_comparativo_cfop(
            writer, bl_c, bl_n, "D190", "ONDE_AGIR_CTe"
        )
        if not ok_nfe and not ok_cte:
            return None, (
                "Não encontrei registros C190 (NF-e) nem D190 (CT-e) nos dois SPED. "
                "No .txt, o C190 deve seguir o C100 e o D190 o D100; "
                "no Excel, use abas C100/C190 e D100/D190."
            )
        _xlsx_aplicar_abas_rosa(writer)
    output.seek(0)
    output = _xlsx_reordenar_abas_por_lista(output, _ORDEM_ABAS_COMP_NFE_CTE)
    return output, "Sucesso"


def gerar_excel_cfop_comparativo_dois_speds(
    file_sped_cliente,
    file_sped_nascel,
    nome_arquivo_cliente: str = "",
    nome_arquivo_nascel: str = "",
    fonte_cfop: str = "C190",
):
    """
    Formato tipo relatório comparativo (duas origens): abas separadas para SPED Cliente e SPED Nascel,
    mais tabelas de confronto por CFOP (diferença Cliente−Nascel), não um único ficheiro “somado”.

    Abas conforme fonte_cfop:
    - C190: bloco C (C100, C190 × Cliente/Nascel) + comparativo + ONDE_AGIR.
    - D190: bloco D (D100, D190 × Cliente/Nascel) + comparativo + ONDE_AGIR.

    Os dois SPED são lidos em paralelo (ThreadPoolExecutor) — apenas para acelerar; cada um mantém a sua identidade.

    Para dois ficheiros .xlsx num ZIP com uma única dupla de leituras, use `gerar_zip_cfop_c190_e_d190_dois_speds`.
    """
    if fonte_cfop not in ("C190", "D190"):
        return None, "fonte_cfop deve ser C190 ou D190."

    try:
        bl_c, bl_n = _carregar_dois_speds_em_paralelo(
            file_sped_cliente,
            file_sped_nascel,
            nome_arquivo_cliente,
            nome_arquivo_nascel,
        )
        return _gerar_excel_cfop_comparativo_de_blocos(bl_c, bl_n, fonte_cfop)
    except Exception as e:
        return None, f"Erro: {str(e)}"


def gerar_excel_cfop_comparativo_nfe_e_cte_dois_speds(
    file_sped_cliente,
    file_sped_nascel,
    nome_arquivo_cliente: str = "",
    nome_arquivo_nascel: str = "",
):
    """
    Dois SPED em paralelo: **um** Excel com comparativo **NF-e (C190)** e **CT-e (D190)** —
    abas Cliente/Nascel + Comparativo_CFOP_C190 + ONDE_AGIR_NF_e, e bloco D100/D190 +
    Comparativo_CFOP_D190 + ONDE_AGIR_CTe (cada bloco só entra se houver linhas).
    """
    try:
        bl_c, bl_n = _carregar_dois_speds_em_paralelo(
            file_sped_cliente,
            file_sped_nascel,
            nome_arquivo_cliente,
            nome_arquivo_nascel,
        )
        return _gerar_excel_cfop_comparativo_nfe_cte_de_blocos(bl_c, bl_n)
    except Exception as e:
        return None, f"Erro: {str(e)}"


def gerar_zip_cfop_c190_e_d190_dois_speds(
    file_sped_cliente,
    file_sped_nascel,
    nome_arquivo_cliente: str = "",
    nome_arquivo_nascel: str = "",
):
    """
    Mesmos dois ficheiros Excel que gerar C190 e D190 à parte, mas com uma única leitura
    paralela Cliente + Nascel (evita parse duplicado dos .txt).
    """
    try:
        bl_c, bl_n = _carregar_dois_speds_em_paralelo(
            file_sped_cliente,
            file_sped_nascel,
            nome_arquivo_cliente,
            nome_arquivo_nascel,
        )
        bio_c, msg_c = _gerar_excel_cfop_comparativo_de_blocos(
            bl_c, bl_n, "C190"
        )
        bio_d, msg_d = _gerar_excel_cfop_comparativo_de_blocos(
            bl_c, bl_n, "D190"
        )
        if bio_c and bio_d:
            zbuf = io.BytesIO()
            with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr(
                    "Detetive_CFOP_C190_2SPED.xlsx",
                    bio_c.getvalue(),
                )
                zf.writestr(
                    "Detetive_CFOP_D190_2SPED.xlsx",
                    bio_d.getvalue(),
                )
            zbuf.seek(0)
            return zbuf, "Sucesso"
        err = " ".join(
            x
            for x in (
                (msg_c or "") if not bio_c else "",
                (msg_d or "") if not bio_d else "",
            )
            if x
        )
        return None, err or "Não foi possível gerar um ou ambos os relatórios."
    except Exception as e:
        return None, f"Erro: {str(e)}"


def gerar_excel_cfop_um_sped(
    file_sped,
    nome_arquivo: str = "",
    fonte_cfop: str = "C190",
    rotulo_lado: str = "",
):
    """
    Um único arquivo SPED: mesma «tabelinha» somatório por CFOP (estilo comparativo Tabela 1)
    e abas C100/C190 ou D100/D190 com coluna LADO. Sem Tabela 2/3 e sem ONDE_AGIR.
    """
    if fonte_cfop not in ("C190", "D190"):
        return None, "fonte_cfop deve ser C190 ou D190."
    if not rotulo_lado:
        rotulo_lado = ROTULO_CLIENTE
    try:
        if hasattr(file_sped, "seek"):
            file_sped.seek(0)
        bl = carregar_blocos_sped_completos(file_sped, nome_arquivo)
        if fonte_cfop == "C190":
            bl["C190"] = _garantir_chv_em_c190(bl["C100"], bl["C190"])
        else:
            bl["D190"] = _garantir_chv_em_d190(bl["D100"], bl["D190"])

        chave_df = "D190" if fonte_cfop == "D190" else "C190"
        df = bl[chave_df]

        if df.empty:
            if fonte_cfop == "D190":
                return None, (
                    "Não encontrei registros D190. No .txt, cada D190 deve vir após o D100 (CT-e); "
                    "no Excel, use abas D100 e D190."
                )
            return None, (
                "Não encontrei registros C190. No .txt, o C190 deve vir logo após o C100; "
                "no Excel, use abas C100 e C190 com colunas padrão EFD."
            )

        tab = tabela_somatorio_cfop_estilo_excel(df)
        c100_x = _df_c100_planilha(_df_com_coluna_lado(bl["C100"], rotulo_lado))
        d100_x = _df_c100_planilha(_df_com_coluna_lado(bl["D100"], rotulo_lado))
        c190_x = _df_export_c190_planilha(bl["C190"], rotulo_lado)
        d190_x = _df_export_d190_planilha(bl["D190"], rotulo_lado)

        nome_comp = (
            "Somatorio_CFOP_C190_um_SPED"
            if fonte_cfop == "C190"
            else "Somatorio_CFOP_D190_um_SPED"
        )
        sh_c100, sh_c190 = "C100", "C190"
        sh_d100, sh_d190 = "D100", "D190"

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            if fonte_cfop == "C190":
                c100_x.to_excel(writer, index=False, sheet_name=sh_c100)
                c190_x.to_excel(writer, index=False, sheet_name=sh_c190)
                _xlsx_reforcar_detalhe_texto_e_numero(
                    writer, sh_c190, c190_x, _COL_DETALHE_C190
                )
                abas_largura = (sh_c100, sh_c190)
                df_det = c190_x
                sh_det = sh_c190
            else:
                d100_x.to_excel(writer, index=False, sheet_name=sh_d100)
                d190_x.to_excel(writer, index=False, sheet_name=sh_d190)
                _xlsx_reforcar_detalhe_texto_e_numero(
                    writer, sh_d190, d190_x, _COL_DETALHE_D190
                )
                abas_largura = (sh_d100, sh_d190)
                df_det = d190_x
                sh_det = sh_d190

            _escrever_somatorio_cfop_unica(
                writer,
                tab,
                nome_comp,
                fonte_cfop,
                rotulo_lado,
                sh_det,
                df_det,
            )

            for name in abas_largura:
                if name in writer.sheets:
                    writer.sheets[name].set_column(0, 40, 14)
            if fonte_cfop == "C190":
                _xlsx_largura_coluna_chave(writer, sh_c100, c100_x, _COL_C100_PLAN)
                _xlsx_largura_coluna_chave(writer, sh_c190, c190_x, _COL_DETALHE_C190)
            else:
                _xlsx_largura_coluna_chave(writer, sh_d100, d100_x, _COL_C100_PLAN)
                _xlsx_largura_coluna_chave(writer, sh_d190, d190_x, _COL_DETALHE_D190)

            _xlsx_aplicar_abas_rosa(writer)
        output.seek(0)
        return output, "Sucesso"
    except Exception as e:
        return None, f"Erro: {str(e)}"


def gerar_excel_cfop_um_sped_completo(
    file_sped,
    nome_arquivo: str = "",
    rotulo_lado: str = "",
):
    """
    Um único SPED: abas C100/C190 + somatório CFOP (C190) e D100/D190 + somatório CFOP (D190)
    no mesmo Excel, para cada bloco que tiver linhas. Sem comparativo entre dois ficheiros.
    """
    if not rotulo_lado:
        rotulo_lado = ROTULO_CLIENTE
    try:
        if hasattr(file_sped, "seek"):
            file_sped.seek(0)
        bl = carregar_blocos_sped_completos(file_sped, nome_arquivo)
        bl["C190"] = _garantir_chv_em_c190(bl["C100"], bl["C190"])
        bl["D190"] = _garantir_chv_em_d190(bl["D100"], bl["D190"])

        tem_c = not bl["C190"].empty
        tem_d = not bl["D190"].empty
        if not tem_c and not tem_d:
            return None, (
                "Não encontrei registros C190 nem D190. No .txt, o C190 deve seguir o C100 e o D190 o D100; "
                "no Excel, use abas C100/C190 e D100/D190."
            )

        c100_x = _df_c100_planilha(_df_com_coluna_lado(bl["C100"], rotulo_lado))
        d100_x = _df_c100_planilha(_df_com_coluna_lado(bl["D100"], rotulo_lado))
        c190_x = _df_export_c190_planilha(bl["C190"], rotulo_lado)
        d190_x = _df_export_d190_planilha(bl["D190"], rotulo_lado)

        sh_c100, sh_c190 = "C100", "C190"
        sh_d100, sh_d190 = "D100", "D190"
        nome_comp_c = "Somatorio_CFOP_C190_um_SPED"
        nome_comp_d = "Somatorio_CFOP_D190_um_SPED"

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            if tem_c:
                tab_c = tabela_somatorio_cfop_estilo_excel(bl["C190"])
                c100_x.to_excel(writer, index=False, sheet_name=sh_c100)
                c190_x.to_excel(writer, index=False, sheet_name=sh_c190)
                _xlsx_reforcar_detalhe_texto_e_numero(
                    writer, sh_c190, c190_x, _COL_DETALHE_C190
                )
                if len(c100_x) > 0:
                    _xlsx_tabela_sobre_df_escrito(writer.sheets[sh_c100], c100_x)
                if len(c190_x) > 0:
                    _xlsx_tabela_sobre_df_escrito(writer.sheets[sh_c190], c190_x)
                _escrever_somatorio_cfop_unica(
                    writer,
                    tab_c,
                    nome_comp_c,
                    "C190",
                    rotulo_lado,
                    sh_c190,
                    c190_x,
                )
            if tem_d:
                tab_d = tabela_somatorio_cfop_estilo_excel(bl["D190"])
                d100_x.to_excel(writer, index=False, sheet_name=sh_d100)
                d190_x.to_excel(writer, index=False, sheet_name=sh_d190)
                _xlsx_reforcar_detalhe_texto_e_numero(
                    writer, sh_d190, d190_x, _COL_DETALHE_D190
                )
                if len(d100_x) > 0:
                    _xlsx_tabela_sobre_df_escrito(writer.sheets[sh_d100], d100_x)
                if len(d190_x) > 0:
                    _xlsx_tabela_sobre_df_escrito(writer.sheets[sh_d190], d190_x)
                _escrever_somatorio_cfop_unica(
                    writer,
                    tab_d,
                    nome_comp_d,
                    "D190",
                    rotulo_lado,
                    sh_d190,
                    d190_x,
                )

            for name in writer.sheets:
                writer.sheets[name].set_column(0, 40, 14)
            if tem_c:
                _xlsx_largura_coluna_chave(writer, sh_c100, c100_x, _COL_C100_PLAN)
                _xlsx_largura_coluna_chave(writer, sh_c190, c190_x, _COL_DETALHE_C190)
            if tem_d:
                _xlsx_largura_coluna_chave(writer, sh_d100, d100_x, _COL_C100_PLAN)
                _xlsx_largura_coluna_chave(writer, sh_d190, d190_x, _COL_DETALHE_D190)

            _xlsx_aplicar_abas_rosa(writer)
        output.seek(0)
        return output, "Sucesso"
    except Exception as e:
        return None, f"Erro: {str(e)}"


def _reg_id_de_nome_aba(nome: str) -> str:
    nome = (nome or "").strip()
    if not nome:
        return nome
    parte = re.split(r"\s*-\s*", nome, maxsplit=1)[0].strip()
    return parte.upper()


def _decodificar_txt(file_obj) -> str:
    file_obj.seek(0)
    raw = file_obj.read()
    if isinstance(raw, str):
        return raw
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace")


def _partes_linha_sped(line: str) -> list[str] | None:
    line = line.strip()
    if not line or line.startswith("#"):
        return None
    parts = line.split("|")
    if parts and parts[0] == "":
        parts = parts[1:]
    if parts and parts[-1] == "":
        parts = parts[:-1]
    if not parts or not parts[0].strip():
        return None
    return parts


def _reforcar_valores_sped_txt_por_posicao(texto: str, blocos: dict[str, pd.DataFrame]) -> None:
    """
    Sobrescreve VL_DOC / totais C190 (e D100/D190) com o split | posicional do Guia.
    O EFDReader fixa N colunas e trunca linhas; arquivos com campos a mais/menos deslocam
    VL_* — o parser por posição replica o comportamento típico da ferramenta de referência.
    """
    c100_vl: list[str] = []
    c190_src: list[dict[str, str]] = []
    d100_vl: list[str] = []
    d190_src: list[dict[str, str]] = []

    for line in texto.splitlines():
        parts = _partes_linha_sped(line)
        if not parts:
            continue
        reg = parts[0].strip().upper()
        if reg == "C100":
            c100_vl.append(parts[11] if len(parts) > 11 else "")
        elif reg == "C190":
            c190_src.append(
                {
                    "VL_OPR": parts[4] if len(parts) > 4 else "",
                    "VL_BC_ICMS": parts[5] if len(parts) > 5 else "",
                    "VL_ICMS": parts[6] if len(parts) > 6 else "",
                    "VL_BC_ICMS_ST": parts[7] if len(parts) > 7 else "",
                    "VL_ICMS_ST": parts[8] if len(parts) > 8 else "",
                    "VL_IPI": parts[10] if len(parts) > 10 else "",
                }
            )
        elif reg == "D100":
            d100_vl.append(parts[14] if len(parts) > 14 else "")
        elif reg == "D190":
            d190_src.append(
                {
                    "VL_OPR": parts[4] if len(parts) > 4 else "",
                    "VL_BC_ICMS": parts[5] if len(parts) > 5 else "",
                    "VL_ICMS": parts[6] if len(parts) > 6 else "",
                    "VL_RED_BC": parts[7] if len(parts) > 7 else "",
                }
            )

    def _aplica_col(df: pd.DataFrame, valores: list[str], col: str) -> pd.DataFrame:
        if df.empty or col not in df.columns or not valores:
            return df
        n = min(len(df), len(valores))
        out = df.copy()
        j = out.columns.get_loc(col)
        if isinstance(j, slice):
            return df
        for i in range(n):
            out.iat[i, j] = valores[i]
        return out

    def _aplica_map(df: pd.DataFrame, linhas: list[dict[str, str]]) -> pd.DataFrame:
        if df.empty or not linhas:
            return df
        n = min(len(df), len(linhas))
        out = df.copy()
        for i in range(n):
            for col, v in linhas[i].items():
                if col not in out.columns:
                    continue
                j = out.columns.get_loc(col)
                if isinstance(j, slice):
                    continue
                out.iat[i, j] = v
        return out

    if "C100" in blocos:
        blocos["C100"] = _aplica_col(blocos["C100"], c100_vl, "VL_DOC")
    if "C190" in blocos:
        blocos["C190"] = _aplica_map(blocos["C190"], c190_src)
    if "D100" in blocos:
        blocos["D100"] = _aplica_col(blocos["D100"], d100_vl, "VL_DOC")
    if "D190" in blocos:
        blocos["D190"] = _aplica_map(blocos["D190"], d190_src)


def _listar_regs_txt(file_obj) -> list[str]:
    text = _decodificar_txt(file_obj)
    regs = set()
    for line in text.splitlines():
        parts = _partes_linha_sped(line)
        if parts:
            regs.add(parts[0].strip().upper())
    file_obj.seek(0)
    return sorted(regs)


def _parse_sped_txt(file_obj) -> dict[str, pd.DataFrame]:
    text = _decodificar_txt(file_obj)
    buckets: dict[str, list[dict]] = {}
    for line in text.splitlines():
        parts = _partes_linha_sped(line)
        if not parts:
            continue
        reg = parts[0].strip().upper()
        row = {"REG": reg}
        for i, val in enumerate(parts[1:], start=2):
            row[f"C{i:02d}"] = val
        buckets.setdefault(reg, []).append(row)
    out: dict[str, pd.DataFrame] = {}
    for reg, rows in buckets.items():
        df = pd.DataFrame(rows)
        df = df.fillna("").astype(str)
        out[reg] = _aplicar_aliases_reg(reg, df)
    file_obj.seek(0)
    return out


def _aplicar_aliases_reg(reg: str, df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for src, dst in ALIASES_POR_REG.get(reg, []):
        if src in df.columns and dst not in df.columns:
            df[dst] = df[src]
    return df


def _listar_abas_excel(file_obj):
    file_obj.seek(0)
    nomes = pd.ExcelFile(file_obj).sheet_names
    file_obj.seek(0)
    return nomes


def listar_registros_arquivo(file_obj, filename: str) -> list[str]:
    fn = (filename or "").lower()
    if fn.endswith(".txt"):
        return _listar_regs_txt(file_obj)
    return sorted({_reg_id_de_nome_aba(s) for s in _listar_abas_excel(file_obj)})


def _carregar_mapa_abas(file_obj, filename: str) -> dict[str, pd.DataFrame]:
    fn = (filename or "").lower()
    if fn.endswith(".txt"):
        return _parse_sped_txt(file_obj)
    file_obj.seek(0)
    xl = pd.ExcelFile(file_obj)
    out: dict[str, pd.DataFrame] = {}
    for sh in xl.sheet_names:
        key = _reg_id_de_nome_aba(sh)
        file_obj.seek(0)
        df = pd.read_excel(file_obj, sheet_name=sh, dtype=str)
        df = _aplicar_aliases_reg(key, df)
        if key in out:
            out[key] = pd.concat([out[key], df], ignore_index=True)
        else:
            out[key] = df
    file_obj.seek(0)
    return out


def _normalizar_coluna_chave(df: pd.DataFrame, chave: str) -> pd.DataFrame:
    df = df.copy()
    s = df[chave].astype(str).str.strip()
    if chave == "CNPJ":
        s = s.str.replace(r"\D", "", regex=True)
    df[chave] = s
    return df


def _detectar_chave(df_cliente: pd.DataFrame, df_nascel: pd.DataFrame):
    c1, c2 = set(df_cliente.columns), set(df_nascel.columns)
    for k in CHAVES_CANDIDATAS:
        if k in c1 and k in c2:
            return k
    return None


def _fingerprint_linha(row, common_cols):
    return tuple(_norm_val(row[c]) for c in common_cols)


def _comparar_por_chave(df_cliente, df_nascel, chave: str):
    df_cliente = _normalizar_coluna_chave(df_cliente, chave)
    df_nascel = _normalizar_coluna_chave(df_nascel, chave)

    k_cliente = set(df_cliente[chave].dropna())
    k_cliente.discard("")
    k_nascel = set(df_nascel[chave].dropna())
    k_nascel.discard("")

    so_cliente = df_cliente[~df_cliente[chave].isin(k_nascel)].copy()
    so_cliente["TIPO_DIFERENCA"] = f"Só no {ROTULO_CLIENTE}"

    so_nascel = df_nascel[~df_nascel[chave].isin(k_cliente)].copy()
    so_nascel["TIPO_DIFERENCA"] = f"Só no {ROTULO_NASCEL}"

    dup_cliente = df_cliente[df_cliente.duplicated(subset=[chave], keep=False)]
    dup_nascel = df_nascel[df_nascel.duplicated(subset=[chave], keep=False)]

    d_cliente = df_cliente.drop_duplicates(subset=[chave], keep="first")
    d_nascel = df_nascel.drop_duplicates(subset=[chave], keep="first")

    common_cols = sorted(set(d_cliente.columns) & set(d_nascel.columns) - {chave})
    merged = d_cliente.merge(
        d_nascel, on=chave, how="inner", suffixes=(SUF_CLIENTE, SUF_NASCEL)
    )

    def divergencias_por_linha(row):
        difs = []
        for col in common_cols:
            c_cli, c_nas = f"{col}{SUF_CLIENTE}", f"{col}{SUF_NASCEL}"
            if c_cli not in row.index and c_nas not in row.index:
                continue
            if c_cli not in row.index:
                difs.append(f"{col} (só no {ROTULO_CLIENTE})")
                continue
            if c_nas not in row.index:
                difs.append(f"{col} (só no {ROTULO_NASCEL})")
                continue
            if _norm_val(row[c_cli]) != _norm_val(row[c_nas]):
                difs.append(col)
        return ", ".join(difs)

    merged["COLUNAS_DIVERGENTES"] = merged.apply(divergencias_por_linha, axis=1)
    com_divergencia = merged[merged["COLUNAS_DIVERGENTES"] != ""].copy()
    identicos = len(merged) - len(com_divergencia)

    return {
        "modo": f"chave:{chave}",
        "so_cliente": so_cliente,
        "so_nascel": so_nascel,
        "com_divergencia": com_divergencia,
        "dup_cliente": dup_cliente,
        "dup_nascel": dup_nascel,
        "n_so_cliente": len(so_cliente),
        "n_so_nascel": len(so_nascel),
        "n_diverg": len(com_divergencia),
        "n_identicos": identicos,
        "n_dup_cliente": len(dup_cliente),
        "n_dup_nascel": len(dup_nascel),
    }


def _comparar_por_linha(df_cliente, df_nascel):
    common_cols = sorted(set(df_cliente.columns) & set(df_nascel.columns))
    if not common_cols:
        return {
            "modo": "linha (sem colunas comuns)",
            "so_cliente": pd.DataFrame(),
            "so_nascel": pd.DataFrame(),
            "com_divergencia": pd.DataFrame(),
            "dup_cliente": pd.DataFrame(),
            "dup_nascel": pd.DataFrame(),
            "n_so_cliente": 0,
            "n_so_nascel": 0,
            "n_diverg": 0,
            "n_identicos": 0,
            "n_dup_cliente": 0,
            "n_dup_nascel": 0,
            "obs": "Nenhuma coluna comum nesta aba.",
        }

    def fp_series(df):
        return df.apply(lambda r: _fingerprint_linha(r, common_cols), axis=1)

    fp_c = fp_series(df_cliente)
    fp_n = fp_series(df_nascel)

    dup_cliente = df_cliente[fp_c.duplicated(keep=False)].copy()
    dup_nascel = df_nascel[fp_n.duplicated(keep=False)].copy()

    ca = Counter(fp_c.tolist())
    cb = Counter(fp_n.tolist())
    diff_c = ca - cb
    diff_n = cb - ca
    n_pareados = sum((ca & cb).values())

    rem_c = diff_c.copy()
    so_cliente_rows = []
    for idx, row in df_cliente.iterrows():
        fp = fp_c.loc[idx]
        if rem_c[fp] > 0:
            rem_c[fp] -= 1
            so_cliente_rows.append(row)
    so_cliente = (
        pd.DataFrame(so_cliente_rows)
        if so_cliente_rows
        else pd.DataFrame(columns=df_cliente.columns)
    )
    if not so_cliente.empty:
        so_cliente = so_cliente.copy()
        so_cliente["TIPO_DIFERENCA"] = f"Só no {ROTULO_CLIENTE} (linha inteira)"

    rem_n = diff_n.copy()
    so_nascel_rows = []
    for idx, row in df_nascel.iterrows():
        fp = fp_n.loc[idx]
        if rem_n[fp] > 0:
            rem_n[fp] -= 1
            so_nascel_rows.append(row)
    so_nascel = (
        pd.DataFrame(so_nascel_rows)
        if so_nascel_rows
        else pd.DataFrame(columns=df_nascel.columns)
    )
    if not so_nascel.empty:
        so_nascel = so_nascel.copy()
        so_nascel["TIPO_DIFERENCA"] = f"Só no {ROTULO_NASCEL} (linha inteira)"

    return {
        "modo": "linha (todas as colunas comuns)",
        "so_cliente": so_cliente,
        "so_nascel": so_nascel,
        "com_divergencia": pd.DataFrame(),
        "dup_cliente": dup_cliente,
        "dup_nascel": dup_nascel,
        "n_so_cliente": len(so_cliente),
        "n_so_nascel": len(so_nascel),
        "n_diverg": 0,
        "n_identicos": n_pareados,
        "n_dup_cliente": len(dup_cliente),
        "n_dup_nascel": len(dup_nascel),
        "obs": "Sem chave NF-e/CT-e/CNPJ/COD_PART: compara o conteúdo das colunas comuns (linhas idênticas se cancelam); o excedente aparece como só Cliente ou só Nascel.",
    }


def _comparar_uma_aba(df_cliente, df_nascel, nome_aba: str):
    if df_cliente.empty and df_nascel.empty:
        return {
            "aba": nome_aba,
            "modo": "—",
            "so_cliente": pd.DataFrame(),
            "so_nascel": pd.DataFrame(),
            "com_divergencia": pd.DataFrame(),
            "dup_cliente": pd.DataFrame(),
            "dup_nascel": pd.DataFrame(),
            "n_so_cliente": 0,
            "n_so_nascel": 0,
            "n_diverg": 0,
            "n_identicos": 0,
            "n_dup_cliente": 0,
            "n_dup_nascel": 0,
            "obs": "Aba vazia nos dois arquivos.",
        }

    chave = _detectar_chave(df_cliente, df_nascel)
    if chave:
        r = _comparar_por_chave(df_cliente, df_nascel, chave)
        r["aba"] = nome_aba
        r["obs"] = ""
        return r

    r = _comparar_por_linha(df_cliente, df_nascel)
    r["aba"] = nome_aba
    return r


def comparar_dois_sped(
    file_sped_cliente,
    file_sped_nascel,
    abas_selecionadas: list,
    nome_arquivo_cliente: str = "",
    nome_arquivo_nascel: str = "",
):
    try:
        map_c = _carregar_mapa_abas(file_sped_cliente, nome_arquivo_cliente)
        map_n = _carregar_mapa_abas(file_sped_nascel, nome_arquivo_nascel)

        comuns = sorted(set(map_c.keys()) & set(map_n.keys()))
        comuns = [a for a in comuns if a in abas_selecionadas]
        if not comuns:
            return None, "Nenhum registro em comum (ou nenhum selecionado)."

        linhas_resumo = []
        all_so_c, all_so_n, all_div, all_dup_c, all_dup_n = [], [], [], [], []

        for aba in comuns:
            df_c = map_c[aba]
            df_n = map_n[aba]
            r = _comparar_uma_aba(df_c, df_n, aba)

            if not r["so_cliente"].empty:
                t = r["so_cliente"].copy()
                t.insert(0, "ABA_SPED", aba)
                all_so_c.append(t)
            if not r["so_nascel"].empty:
                t = r["so_nascel"].copy()
                t.insert(0, "ABA_SPED", aba)
                all_so_n.append(t)
            if not r["com_divergencia"].empty:
                t = r["com_divergencia"].copy()
                t.insert(0, "ABA_SPED", aba)
                all_div.append(t)
            if not r["dup_cliente"].empty:
                t = r["dup_cliente"].copy()
                t.insert(0, "ABA_SPED", aba)
                all_dup_c.append(t)
            if not r["dup_nascel"].empty:
                t = r["dup_nascel"].copy()
                t.insert(0, "ABA_SPED", aba)
                all_dup_n.append(t)

            linhas_resumo.append(
                {
                    "ABA_SPED": aba,
                    "MODO_COMPARACAO": r["modo"],
                    f"Só_{ROTULO_CLIENTE.replace(' ', '_')}": r["n_so_cliente"],
                    f"Só_{ROTULO_NASCEL.replace(' ', '_')}": r["n_so_nascel"],
                    "Com_divergencia_campos": r["n_diverg"],
                    "Pareados_ou_identicos": r["n_identicos"],
                    f"Dup_{ROTULO_CLIENTE.replace(' ', '_')}": r["n_dup_cliente"],
                    f"Dup_{ROTULO_NASCEL.replace(' ', '_')}": r["n_dup_nascel"],
                    "Observacao": r.get("obs", ""),
                }
            )

        resumo = pd.DataFrame(linhas_resumo)
        totais = pd.DataFrame(
            [
                {
                    "Indicador": "Total de registros (blocos) comparados",
                    "Quantidade": len(comuns),
                },
                {
                    "Indicador": f"Soma linhas só {ROTULO_CLIENTE}",
                    "Quantidade": int(resumo[f"Só_{ROTULO_CLIENTE.replace(' ', '_')}"].sum()),
                },
                {
                    "Indicador": f"Soma linhas só {ROTULO_NASCEL}",
                    "Quantidade": int(resumo[f"Só_{ROTULO_NASCEL.replace(' ', '_')}"].sum()),
                },
                {
                    "Indicador": "Soma linhas com divergência de campos (modo chave)",
                    "Quantidade": int(resumo["Com_divergencia_campos"].sum()),
                },
            ]
        )

        def _concat_or_empty(parts, cols_msg):
            if not parts:
                return pd.DataFrame(columns=["ABA_SPED"] if cols_msg else [])
            return pd.concat(parts, ignore_index=True)

        out_so_c = _concat_or_empty(all_so_c, True)
        out_so_n = _concat_or_empty(all_so_n, True)
        out_div = _concat_or_empty(all_div, True)
        out_dup_c = _concat_or_empty(all_dup_c, True)
        out_dup_n = _concat_or_empty(all_dup_n, True)

        onde_agir = montar_onde_agir_sped(out_div, out_so_c, out_so_n)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            onde_agir.to_excel(writer, index=False, sheet_name="ONDE_AGIR")
            totais.to_excel(writer, index=False, sheet_name="Resumo_Geral")
            resumo.to_excel(writer, index=False, sheet_name="Resumo_por_aba")
            out_so_c.to_excel(writer, index=False, sheet_name="So_Cliente")
            out_so_n.to_excel(writer, index=False, sheet_name="So_Nascel")
            out_div.to_excel(writer, index=False, sheet_name="Campos_divergentes")
            if not out_dup_c.empty:
                out_dup_c.to_excel(writer, index=False, sheet_name="Dup_Cliente")
            if not out_dup_n.empty:
                out_dup_n.to_excel(writer, index=False, sheet_name="Dup_Nascel")

            if "ONDE_AGIR" in writer.sheets:
                w_oa = writer.sheets["ONDE_AGIR"]
                w_oa.set_column(0, 0, 6)
                w_oa.set_column(1, 3, 16)
                w_oa.set_column(4, 4, 88)

            for name, sheet in writer.sheets.items():
                if name == "ONDE_AGIR":
                    continue
                sheet.set_column(0, 60, 16)

            _xlsx_aplicar_abas_rosa(writer)
        output.seek(0)
        return output, "Sucesso"
    except Exception as e:
        return None, f"Erro: {str(e)}"
